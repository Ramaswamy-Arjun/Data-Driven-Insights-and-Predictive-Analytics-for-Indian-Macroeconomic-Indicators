from openpyxl import load_workbook

# Load the Excel workbook
file_path = 'E:\GST Collections\Tax-Collection-on-GST-Portal-2023-2024.xlsx'  # Update with your file path
wb = load_workbook(filename=file_path)

# Access Sheet 1 and Sheet 2
sheet1 = wb['collection_2023-24']  # Assuming the name of sheet 1 is 'Sheet1'
if "Sheet1" not in wb.sheetnames:
        sheet2 = wb.create_sheet("Sheet1")
else:
        sheet2 = wb["Sheet1"]

# Loop through the rows from 7 to 47 in Sheet1
for i, row in enumerate(range(5, 46), start=1):  # Start enumerating from 1 for Sheet2
    for col in range(1, sheet1.max_column + 1):  # Loop through all columns
        # Get value from Sheet 1
        cell_value = sheet1.cell(row=row, column=col).value
        
        # Paste value into Sheet 2, starting from row 1
        sheet2.cell(row=i, column=col).value = cell_value

# Save the workbook after copying
wb.save(filename='Tax-Collection-on-GST-Portal-2023-2024-updated.xlsx')
